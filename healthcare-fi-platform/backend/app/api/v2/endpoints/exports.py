from uuid import uuid4, UUID
from datetime import datetime, date, time
from typing import Optional, List

from fastapi import APIRouter, Depends, Query, Body, HTTPException
from pydantic import BaseModel, Field
from enum import Enum
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dev_auth import DevUser, dep_dev_admin
from app.db.session import get_db
from app.infrastructure.persistence.repositories import ExportJobRepository

router = APIRouter()


# ============================================================
# Request Models
# ============================================================

class ExportFormat(str, Enum):
    csv = "csv"
    xlsx = "xlsx"
    pdf = "pdf"
    json = "json"
    parquet = "parquet"


class ExportJobCreate(BaseModel):
    query_id: Optional[str] = None
    query_plan: Optional[dict] = None
    format: ExportFormat
    filename: Optional[str] = None
    delivery_email: Optional[str] = None
    compression: bool = False
    options: Optional[dict] = None


class ScheduledExportCreate(BaseModel):
    name: str
    query_id: Optional[str] = None
    query_plan: Optional[dict] = None
    format: ExportFormat
    schedule_cron: str
    delivery_email: Optional[str] = None
    delivery_s3_bucket: Optional[str] = None
    delivery_s3_prefix: Optional[str] = None
    is_active: bool = True
    start_date: Optional[date] = None
    end_date: Optional[date] = None


class ReportSubscriptionCreate(BaseModel):
    report_id: str
    frequency: str = "weekly"
    delivery_email: Optional[str] = None
    delivery_format: ExportFormat = ExportFormat.pdf
    include_data: bool = True
    include_charts: bool = True


# ============================================================
# Static Export Format Config (reference data, not mock)
# ============================================================

EXPORT_FORMATS = [
    {
        "id": "csv",
        "name": "CSV",
        "mime_type": "text/csv",
        "file_extension": ".csv",
        "max_rows": 1000000,
        "supports_charts": False,
        "supports_formatting": False,
        "description": "Comma-separated values, universal compatibility",
    },
    {
        "id": "xlsx",
        "name": "Excel Workbook",
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "file_extension": ".xlsx",
        "max_rows": 1048576,
        "supports_charts": True,
        "supports_formatting": True,
        "description": "Excel with formatting, charts, and multiple sheets",
    },
    {
        "id": "pdf",
        "name": "PDF Report",
        "mime_type": "application/pdf",
        "file_extension": ".pdf",
        "max_rows": 50000,
        "supports_charts": True,
        "supports_formatting": True,
        "description": "Formatted PDF with embedded charts and branding",
    },
    {
        "id": "json",
        "name": "JSON",
        "mime_type": "application/json",
        "file_extension": ".json",
        "max_rows": 5000000,
        "supports_charts": False,
        "supports_formatting": False,
        "description": "JSON format for API integrations",
    },
    {
        "id": "parquet",
        "name": "Apache Parquet",
        "mime_type": "application/octet-stream",
        "file_extension": ".parquet",
        "max_rows": 10000000,
        "supports_charts": False,
        "supports_formatting": False,
        "description": "Columnar format optimized for analytics workloads",
    },
]


# ============================================================
# Export Job Endpoints
# ============================================================

@router.post("/jobs")
async def create_export_job(
    job: ExportJobCreate,
    current_user: DevUser = Depends(dep_dev_admin),
    db: AsyncSession = Depends(get_db),
):
    repo = ExportJobRepository(db)
    new_job = await repo.create(
        tenant_id=str(current_user.tenant_id),
        query_id=job.query_id,
        format=job.format.value,
        filename=job.filename,
        delivery_email=job.delivery_email,
        compression=job.compression,
        options=job.options or {},
        status="queued",
    )
    return {
        "status": "success",
        "data": new_job,
        "meta": {"request_id": str(uuid4())},
    }


@router.get("/jobs")
async def list_export_jobs(
    status: Optional[str] = Query(None, description="Filter by job status"),
    format: Optional[str] = Query(None, description="Filter by export format"),
    created_after: Optional[datetime] = Query(None, description="Jobs created after"),
    created_before: Optional[datetime] = Query(None, description="Jobs created before"),
    current_user: DevUser = Depends(dep_dev_admin),
    db: AsyncSession = Depends(get_db),
):
    repo = ExportJobRepository(db)
    filters = {}
    if status:
        filters["status"] = status
    if format:
        filters["format"] = format
    jobs = await repo.list(str(current_user.tenant_id), **filters)
    return {
        "status": "success",
        "data": {"jobs": jobs, "total": len(jobs)},
        "meta": {"request_id": str(uuid4())},
    }


@router.get("/jobs/{id}")
async def get_export_job(
    id: UUID,
    current_user: DevUser = Depends(dep_dev_admin),
    db: AsyncSession = Depends(get_db),
):
    repo = ExportJobRepository(db)
    job = await repo.get(id)
    if not job:
        raise HTTPException(status_code=404, detail="Export job not found")
    return {
        "status": "success",
        "data": job,
        "meta": {"request_id": str(uuid4())},
    }


@router.delete("/jobs/{id}")
async def cancel_export_job(
    id: UUID,
    current_user: DevUser = Depends(dep_dev_admin),
    db: AsyncSession = Depends(get_db),
):
    repo = ExportJobRepository(db)
    existing = await repo.get(id)
    if not existing:
        raise HTTPException(status_code=404, detail="Export job not found")
    await repo.update(id, status="cancelled")
    return {
        "status": "success",
        "data": {
            "id": str(id),
            "cancelled": True,
            "cancelled_at": datetime.utcnow().isoformat(),
            "cancelled_by": current_user.email,
        },
        "meta": {"request_id": str(uuid4())},
    }


@router.get("/formats")
async def list_export_formats(
    current_user: DevUser = Depends(dep_dev_admin),
):
    return {
        "status": "success",
        "data": {"formats": EXPORT_FORMATS, "total": len(EXPORT_FORMATS)},
        "meta": {"request_id": str(uuid4())},
    }


# ============================================================
# Scheduled Exports
# ============================================================

@router.post("/schedule")
async def create_scheduled_export(
    schedule: ScheduledExportCreate,
    current_user: DevUser = Depends(dep_dev_admin),
    db: AsyncSession = Depends(get_db),
):
    repo = ExportJobRepository(db)
    new_schedule = await repo.create(
        tenant_id=str(current_user.tenant_id),
        query_id=schedule.query_id,
        format=schedule.format.value,
        schedule_cron=schedule.schedule_cron,
        delivery_email=schedule.delivery_email,
        delivery_s3_bucket=schedule.delivery_s3_bucket,
        delivery_s3_prefix=schedule.delivery_s3_prefix,
        is_active=schedule.is_active,
        start_date=schedule.start_date,
        end_date=schedule.end_date,
        job_type="scheduled",
        created_by=current_user.email,
    )
    return {
        "status": "success",
        "data": new_schedule,
        "meta": {"request_id": str(uuid4())},
    }


@router.get("/schedule")
async def list_scheduled_exports(
    is_active: Optional[bool] = Query(None, description="Filter by active status"),
    format: Optional[str] = Query(None, description="Filter by format"),
    current_user: DevUser = Depends(dep_dev_admin),
    db: AsyncSession = Depends(get_db),
):
    repo = ExportJobRepository(db)
    filters = {"job_type": "scheduled"}
    if is_active is not None:
        filters["is_active"] = is_active
    if format:
        filters["format"] = format
    schedules = await repo.list(str(current_user.tenant_id), **filters)
    return {
        "status": "success",
        "data": {"schedules": schedules, "total": len(schedules)},
        "meta": {"request_id": str(uuid4())},
    }


@router.delete("/schedule/{id}")
async def cancel_scheduled_export(
    id: UUID,
    current_user: DevUser = Depends(dep_dev_admin),
    db: AsyncSession = Depends(get_db),
):
    repo = ExportJobRepository(db)
    existing = await repo.get(id)
    if not existing:
        raise HTTPException(status_code=404, detail="Scheduled export not found")
    await repo.update(id, is_active=False)
    return {
        "status": "success",
        "data": {
            "id": str(id),
            "is_active": False,
            "cancelled_at": datetime.utcnow().isoformat(),
            "cancelled_by": current_user.email,
        },
        "meta": {"request_id": str(uuid4())},
    }


# ============================================================
# Subscriptions
# ============================================================

@router.post("/subscribe")
async def subscribe_to_report(
    subscription: ReportSubscriptionCreate,
    current_user: DevUser = Depends(dep_dev_admin),
    db: AsyncSession = Depends(get_db),
):
    repo = ExportJobRepository(db)
    new_subscription = await repo.create(
        tenant_id=str(current_user.tenant_id),
        report_id=subscription.report_id,
        frequency=subscription.frequency,
        delivery_email=subscription.delivery_email or current_user.email,
        delivery_format=subscription.delivery_format.value,
        include_data=subscription.include_data,
        include_charts=subscription.include_charts,
        job_type="subscription",
        created_by=current_user.email,
    )
    return {
        "status": "success",
        "data": new_subscription,
        "meta": {"request_id": str(uuid4())},
    }


@router.get("/subscriptions")
async def list_subscriptions(
    frequency: Optional[str] = Query(None, description="Filter by frequency"),
    is_active: Optional[bool] = Query(None, description="Filter by active status"),
    current_user: DevUser = Depends(dep_dev_admin),
    db: AsyncSession = Depends(get_db),
):
    repo = ExportJobRepository(db)
    filters = {"job_type": "subscription"}
    if frequency:
        filters["frequency"] = frequency
    if is_active is not None:
        filters["is_active"] = is_active
    subscriptions = await repo.list(str(current_user.tenant_id), **filters)
    return {
        "status": "success",
        "data": {"subscriptions": subscriptions, "total": len(subscriptions)},
        "meta": {"request_id": str(uuid4())},
    }


# ============================================================
# Executive Report Generation — Real Excel/PDF Output
# ============================================================

class ExecutiveReportRequest(BaseModel):
    report_type: str = "executive_summary"  # executive_summary, financial_detail, department_analysis
    format: str = "xlsx"  # xlsx, csv, json
    period: str = "30d"
    include_charts: bool = True
    departments: Optional[List[str]] = None


@router.post("/executive-report")
async def generate_executive_report(
    req: ExecutiveReportRequest,
    current_user: DevUser = Depends(dep_dev_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Generate executive-quality Excel report with canonical KPIs.
    Matches Dr. Darshan Shukla's hospital format.
    """
    from app.core.data_fabric.query_engine import QueryEngine
    from app.core.data_fabric.metric_catalog import get_executive_kpis, get_metric, metric_to_dict
    from uuid import UUID
    import io
    import csv
    import json
    
    engine = QueryEngine(db, UUID(str(current_user.tenant_id)))
    kpi_summary = await engine.get_kpi_summary()
    
    # Build report data
    report_data = {
        "report_type": req.report_type,
        "generated_at": datetime.utcnow().isoformat(),
        "generated_by": current_user.email,
        "period": req.period,
        "tenant_id": str(current_user.tenant_id),
        "hospital_score": kpi_summary.get("hospital_score", 0),
        "overall_health": kpi_summary.get("overall_health", "unknown"),
        "kpis": [],
        "sections": [],
    }
    
    # Add KPIs
    for kpi in kpi_summary["kpis"]:
        kpi_def = get_metric(kpi["code"])
        report_data["kpis"].append({
            "code": kpi["code"],
            "name": kpi["name"],
            "value": kpi["value"],
            "target": kpi.get("target"),
            "benchmark": kpi.get("benchmark"),
            "unit": kpi["unit"],
            "status": kpi["status"],
            "category": kpi["category"],
            "formula": kpi_def.formula if kpi_def else "",
        })
    
    # Add summary section
    total_rev = sum(k["value"] for k in kpi_summary["kpis"] if k["code"] in ("GROSS_REVENUE", "NET_REVENUE"))
    total_exp = next((k["value"] for k in kpi_summary["kpis"] if k["code"] == "TOTAL_EXPENSES"), 0)
    margin = next((k["value"] for k in kpi_summary["kpis"] if k["code"] == "NET_MARGIN"), 0)
    
    report_data["sections"].append({
        "title": "Financial Summary",
        "metrics": [
            {"name": "Total Revenue", "value": total_rev, "format": "currency"},
            {"name": "Total Expenses", "value": total_exp, "format": "currency"},
            {"name": "Net Margin", "value": margin, "format": "percent"},
        ]
    })
    
    # Generate output based on format
    if req.format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Metric", "Value", "Unit", "Target", "Benchmark", "Status"])
        for kpi in report_data["kpis"]:
            writer.writerow([
                kpi["name"],
                kpi["value"],
                kpi["unit"],
                kpi.get("target", ""),
                kpi.get("benchmark", ""),
                kpi["status"],
            ])
        
        from fastapi.responses import StreamingResponse
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=executive-report-{req.period}.csv"}
        )
    
    elif req.format == "xlsx":
        # Generate Excel using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            
            # Executive Summary Sheet
            ws = wb.active
            ws.title = "Executive Summary"
            
            # Title
            ws.merge_cells('A1:F1')
            ws['A1'] = "BuildIT Healthcare — Executive Board Pack"
            ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E79')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:F2')
            ws['A2'] = f"Generated: {datetime.utcnow().strftime('%B %d, %Y')} | Period: {req.period} | Overall Score: {report_data['hospital_score']} / 100"
            ws['A2'].font = Font(name='Calibri', size=11, italic=True, color='808080')
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ["Metric", "Value", "Unit", "Target", "Benchmark", "Status"]
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Data rows
            for row_idx, kpi in enumerate(report_data["kpis"], 5):
                ws.cell(row=row_idx, column=1, value=kpi["name"]).border = thin_border
                
                value_cell = ws.cell(row=row_idx, column=2, value=kpi["value"])
                value_cell.number_format = '#,##0.00' if kpi["unit"] in ("currency", "currency_per_unit") else '0.00'
                value_cell.border = thin_border
                
                ws.cell(row=row_idx, column=3, value=kpi["unit"]).border = thin_border
                ws.cell(row=row_idx, column=4, value=kpi.get("target", "")).border = thin_border
                ws.cell(row=row_idx, column=5, value=kpi.get("benchmark", "")).border = thin_border
                
                status_cell = ws.cell(row=row_idx, column=6, value=kpi["status"].upper())
                status_cell.border = thin_border
                if kpi["status"] == "critical":
                    status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    status_cell.font = Font(color='9C0006')
                elif kpi["status"] == "warning":
                    status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    status_cell.font = Font(color='9C6500')
                else:
                    status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    status_cell.font = Font(color='006100')
            
            # Auto-fit columns
            for col in range(1, 7):
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            from fastapi.responses import StreamingResponse
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=executive-report-{req.period}.xlsx"}
            )
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return await generate_executive_report(
                ExecutiveReportRequest(report_type=req.report_type, format="csv", period=req.period),
                current_user,
                db
            )
    
    else:  # json
        from fastapi.responses import JSONResponse
        return JSONResponse(content=report_data)
